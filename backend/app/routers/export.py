"""Export calculation results to Excel or CSV."""

import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

router = APIRouter()


@router.post("/export/excel")
async def export_excel(data: dict):
    """
    Export calculation results as an Excel workbook with formatted sheets.
    """
    try:
        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, size=12)
        section_font = Font(bold=True, size=11, color="1F4E79")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws.append(["Power Loss Calculation Report"])
        ws.merge_cells("A1:D1")
        ws["A1"].font = Font(bold=True, size=14)

        ws.append([])
        ws.append(["Device", data.get("module_name", "N/A")])
        ws.append(["Device Type", data.get("device_type", "N/A")])
        ws.append([])

        # Operating conditions
        ws.append(["Operating Conditions"])
        ws.cell(row=ws.max_row, column=1).font = section_font

        conditions = data.get("conditions", {})
        for key, val in conditions.items():
            ws.append([f"  {key}", val])

        ws.append([])
        ws.append(["Loss Summary"])
        ws.cell(row=ws.max_row, column=1).font = section_font

        summary_fields = [
            ("Total Loss (W)", data.get("p_total_loss")),
            ("IGBT Conduction Loss (W)", data.get("p_igbt_cond")),
            ("IGBT Switching Loss (W)", data.get("p_igbt_sw")),
            ("Diode Conduction Loss (W)", data.get("p_diode_cond")),
            ("Diode Switching Loss (W)", data.get("p_diode_sw")),
            ("Brake Loss (W)", data.get("p_brake_loss", 0)),
            ("Efficiency (%)", data.get("efficiency")),
            ("Output Power (W)", data.get("p_out")),
            ("Max Junction Temp (°C)", data.get("t_j_max")),
            ("Max Tj Device", data.get("t_j_max_device")),
            ("Case Temp (°C)", data.get("t_case_est")),
            ("Thermal Iterations", data.get("iteration_count")),
            ("Converged", data.get("converged")),
        ]
        for label, value in summary_fields:
            ws.append([f"  {label}", value])

        # ── Sheet 2: Device Details ───────────────────────────────────
        ws2 = wb.create_sheet("Device Details")
        detail_headers = ["Device", "Type", "P_cond (W)", "P_sw (W)", "P_total (W)", "Tj (°C)"]
        ws2.append(detail_headers)
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        devices = data.get("devices", [])
        for dev in devices:
            ws2.append([
                dev.get("name", ""),
                dev.get("type", ""),
                dev.get("p_cond", 0),
                dev.get("p_sw", 0),
                dev.get("p_total", 0),
                dev.get("t_j", 0),
            ])

        # Summary per type
        ws2.append([])
        ws2.append(["Per-Type Summary"])
        igbt_cond = sum(d["p_cond"] for d in devices if d.get("type") == "IGBT")
        igbt_sw = sum(d["p_sw"] for d in devices if d.get("type") == "IGBT")
        diode_cond = sum(d["p_cond"] for d in devices if d.get("type") == "Diode")
        diode_sw = sum(d["p_sw"] for d in devices if d.get("type") == "Diode")
        ws2.append(["All IGBTs", "", round(igbt_cond, 2), round(igbt_sw, 2),
                     round(igbt_cond + igbt_sw, 2)])
        ws2.append(["All Diodes", "", round(diode_cond, 2), round(diode_sw, 2),
                     round(diode_cond + diode_sw, 2)])

        # ── Sheet 3: Calculation Steps ────────────────────────────────
        ws3 = wb.create_sheet("Calculation Steps")
        steps = data.get("calculation_steps", [])
        for step in steps:
            ws3.append([step.get("title", "")])
            ws3.cell(row=ws3.max_row, column=1).font = section_font
            if step.get("formula"):
                ws3.append([f"Formula: {step['formula']}"])

            step_data = step.get("data", {})
            if isinstance(step_data, dict):
                for key, val in step_data.items():
                    if not isinstance(val, (list, dict)):
                        ws3.append([f"  {key}", val])
            elif isinstance(step_data, list):
                # Thermal iteration log
                if step_data and isinstance(step_data[0], dict):
                    log_headers = list(step_data[0].keys())
                    ws3.append(log_headers)
                    for entry in step_data:
                        ws3.append([entry.get(h, "") for h in log_headers])
                ws3.append([])

        # Auto-width
        for ws in [ws, ws2, ws3]:
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=loss_calculation.xlsx"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export/csv")
async def export_csv(data: dict):
    """
    Export calculation results as a CSV file.
    """
    try:
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["Power Loss Calculation Report"])
        writer.writerow([])
        writer.writerow(["Parameter", "Value"])
        writer.writerow(["Total Loss (W)", data.get("p_total_loss")])
        writer.writerow(["IGBT Conduction Loss (W)", data.get("p_igbt_cond")])
        writer.writerow(["IGBT Switching Loss (W)", data.get("p_igbt_sw")])
        writer.writerow(["Diode Conduction Loss (W)", data.get("p_diode_cond")])
        writer.writerow(["Diode Switching Loss (W)", data.get("p_diode_sw")])
        writer.writerow(["Efficiency (%)", data.get("efficiency")])
        writer.writerow(["Max Tj (°C)", data.get("t_j_max")])
        writer.writerow([])

        writer.writerow(["Device", "Type", "P_cond (W)", "P_sw (W)", "P_total (W)", "Tj (°C)"])
        for dev in data.get("devices", []):
            writer.writerow([
                dev.get("name"), dev.get("type"),
                dev.get("p_cond"), dev.get("p_sw"),
                dev.get("p_total"), dev.get("t_j"),
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=loss_calculation.csv"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
